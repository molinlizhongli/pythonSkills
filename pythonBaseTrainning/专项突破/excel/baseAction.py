# _*_ coding = utf-8 _*_
# @Date : 2021/12/15
# @Time : 10:00
# @NAME ：molin


import openpyxl
from openpyxl.styles import Font,PatternFill,Border,Side,Alignment

'''
    openpyxl支持xlsx,不支持旧版的xls,xls需使用xlrd和xlwt
'''

def excelaction():

    # 获取excel文件
    wb = openpyxl.load_workbook('base.xlsx')
    # 全新excel
    # wbnew = openpyxl.Workbook()
    # 获取第一个表的方法
    sheet = wb.worksheets[0] #法一
    # sheet = wb['区域按天统计'] #法二
    # 没有合并单元格的base.xlsx遍历每一行
    for row in sheet.iter_rows():
        for cell in row:
            print(cell.coordinate, cell.value)
    # 修改数据(3种方式）
    # sheet['A1'].value = '1111'
    # sheet.cell(1,1).value = '1111'
    # sheet.cell(1,1,'1111')

    # 保存
    # wb.save('base.xlsx')

    # Excel的样式（字体，边框等）
    font = Font(name='微软雅黑',size=10,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FF0000')
    sheet['A2'].font = font # 将字体设置到表格A1中
    fill = PatternFill(fill_type='darkUp',start_color='FF0000',end_color='FF0000')
    sheet.cell(2,1).fill = fill # 填充样式
    border = Border(left=Side(border_style='dashDotDot',color='9932CC'),
                    right=Side(border_style='dashDotDot',color='121212'),
                    top=Side(border_style='dashDotDot',color='8B0A50'),
                    bottom=Side(border_style='dashDotDot',color='B3EE3A'))
    sheet.cell(5,4).border = border

    alignment = Alignment(horizontal='center',vertical='center',text_rotation=0,indent=0)
    sheet.cell(5,3).alignment = alignment

    # 超链接
    sheet.cell(6,3).value = '=HYPERLINK("%s","%s")' % ("https://www.badu.com","百度一下")
    sheet['D3'] = '=SUM(A3:E3)'

    # 合并单元格
    sheet.merged_cells('A1:D1')
    # 拆分单元格
    # sheet.unmerge_cells('A1:E1')

    wb.save('base.xlsx')
    #创建一个表
    # wbnew.create_sheet(index=0,title='小猪')
    # wbnew.create_sheet(index=1,title='佩奇')

if __name__ == '__main__':
    excelaction()